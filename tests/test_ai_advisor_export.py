# -*- coding: utf-8 -*-
from datetime import datetime, timedelta
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook

from ai_advisor import api_client, config_snapshot, exporter, history, paths


def _patch_account_dir(monkeypatch, tmp_path):
    monkeypatch.setattr(paths, "account_dir", lambda: str(tmp_path))
    monkeypatch.setattr(paths, "account_id", lambda: "TEST")
    paths.ensure_advisor_dirs()


def _new_history_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    for name, headers in history.SHEETS.items():
        ws = wb.create_sheet(name)
        ws.append(headers)
    return wb


def _row(headers, **values):
    return [values.get(header, "") for header in headers]


def test_advisor_paths_split_history_and_export(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path)

    assert paths.history_path().replace("\\", "/").endswith("advisor_history.xlsx")
    assert paths.history_path().replace("\\", "/").endswith("history/advisor_history.xlsx")
    assert paths.export_path().replace("\\", "/").endswith("advisor/advisor_export.xlsx")
    assert paths.advisor_flow_path().replace("\\", "/").endswith("advisor/advisor_flow.md")
    assert paths.advisor_prompt_path().replace("\\", "/").endswith("advisor/advisor_prompt.md")
    assert paths.advisor_api_settings_path().replace("\\", "/").endswith("advisor_api_settings.json")
    assert not paths.advisor_api_settings_path().replace("\\", "/").endswith("advisor/advisor_api_settings.json")
    assert paths.legacy_advisor_api_settings_path().replace("\\", "/").endswith("advisor/advisor_api_settings.json")
    assert paths.advisor_response_path().replace("\\", "/").endswith("advisor/advisor_response.md")
    assert paths.advisor_response_history_path().replace("\\", "/").endswith(".md")
    assert "/history/advisor_response_" in paths.advisor_response_history_path().replace("\\", "/")
    assert "/history/user_context_" in paths.user_context_history_path().replace("\\", "/")


def test_storage_csv_paths_live_in_account_history(monkeypatch, tmp_path):
    import core.storage_manager as storage_manager

    repo_root = Path.cwd()
    repo_artifact = repo_root / "data" / "TEST_ACCOUNT"
    if repo_artifact.exists():
        import shutil

        shutil.rmtree(repo_artifact)
    monkeypatch.chdir(tmp_path)
    try:
        storage_manager.set_active_account("TEST_ACCOUNT")

        assert storage_manager.MASTER_LOG_FILE.replace("\\", "/").endswith("data/TEST_ACCOUNT/history/trade_history_master.csv")
        assert storage_manager.HISTORY_FILE.replace("\\", "/").endswith("data/TEST_ACCOUNT/history/trade_history_log.csv")
        assert not repo_artifact.exists()
    finally:
        if repo_artifact.exists():
            import shutil

            shutil.rmtree(repo_artifact)


def test_storage_master_csv_gets_full_time_columns(monkeypatch, tmp_path):
    import csv
    import core.storage_manager as storage_manager

    monkeypatch.chdir(tmp_path)
    history_dir = tmp_path / "data" / "TEST_ACCOUNT" / "history"
    history_dir.mkdir(parents=True)
    master = history_dir / "trade_history_master.csv"
    with open(master, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Time", "Ticket", "Symbol", "Type", "Vol", "Entry", "SL", "TP", "Fee", "PnL ($)", "Reason", "Market Mode", "Trigger", "Session_ID", "MAE ($)", "MFE ($)"])
        writer.writerow(["08:34:49 -> 11:12:38", "1", "ETHUSD", "BUY", "1", "1", "1", "1", "0", "10", "Manual_Close", "ANY", "[USER]", "20260609_111238", "0", "10"])

    storage_manager.set_active_account("TEST_ACCOUNT")

    with open(master, "r", newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    header = rows[0]
    row = rows[1]
    assert "Open Time" in header
    assert "Close Time" in header
    assert row[header.index("Open Time")] == "2026-06-09T08:34:49"
    assert row[header.index("Close Time")] == "2026-06-09T11:12:38"


def test_storage_master_csv_keeps_unknown_legacy_close_time_blank(monkeypatch, tmp_path):
    import csv
    import core.storage_manager as storage_manager

    monkeypatch.chdir(tmp_path)
    history_dir = tmp_path / "data" / "TEST_ACCOUNT" / "history"
    history_dir.mkdir(parents=True)
    master = history_dir / "trade_history_master.csv"
    with open(master, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Time", "Ticket", "Symbol", "Type", "Vol", "Entry", "SL", "TP", "Fee", "PnL ($)", "Reason", "Market Mode", "Trigger", "Session_ID", "MAE ($)", "MFE ($)", "Open Time", "Close Time"])
        writer.writerow(["13:52:16 -> 16:16:59", "legacy-session", "ETHUSD", "BUY", "1", "1", "1", "1", "0", "10", "Manual_Close", "ANY", "[LEGACY]", "LEGACY", "0", "10", "", "2026-06-10T10:32:39"])

    storage_manager.set_active_account("TEST_ACCOUNT")

    with open(master, "r", newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    header = rows[0]
    row = rows[1]
    assert row[header.index("Open Time")] == ""
    assert row[header.index("Close Time")] == ""


def test_advisor_folder_has_no_nested_dirs_after_export(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path)
    wb = _new_history_workbook()
    wb.save(paths.history_path())

    result = history.build_export_workbook(export_days=7)

    assert result["ok"] is True
    assert not [p for p in (tmp_path / "advisor").iterdir() if p.is_dir()]


def test_export_workbook_filters_closed_trades_without_touching_full_history(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path)
    now = datetime.now()
    wb = _new_history_workbook()

    closed_headers = history.SHEETS["closed_trades"]
    wb["closed_trades"].append(
        _row(
            closed_headers,
            **{
                "Recorded At": (now - timedelta(days=3)).isoformat(timespec="seconds"),
                "Ticket": "recent",
                "Symbol": "ETHUSD",
                "Exit Time": (now - timedelta(days=3)).isoformat(timespec="seconds"),
                "Profit": "10",
            },
        )
    )
    wb["closed_trades"].append(
        _row(
            closed_headers,
            **{
                "Recorded At": (now - timedelta(days=20)).isoformat(timespec="seconds"),
                "Ticket": "old",
                "Symbol": "ETHUSD",
                "Exit Time": (now - timedelta(days=20)).isoformat(timespec="seconds"),
                "Profit": "-5",
            },
        )
    )

    open_headers = history.SHEETS["open_trades"]
    wb["open_trades"].append(
        _row(open_headers, **{"Recorded At": now.isoformat(timespec="seconds"), "Ticket": "open-1"})
    )
    wb.save(paths.history_path())

    result_7d = history.build_export_workbook(export_days=7)
    assert result_7d["ok"] is True
    assert result_7d["closed_trades"] == 1
    export_wb = load_workbook(paths.export_path())
    assert [export_wb["closed_trades"].cell(r, 2).value for r in range(2, export_wb["closed_trades"].max_row + 1)] == ["recent"]
    assert export_wb["open_trades"].cell(2, 2).value == "open-1"

    result_30d = history.build_export_workbook(export_days=30)
    assert result_30d["closed_trades"] == 2
    export_wb = load_workbook(paths.export_path())
    assert [export_wb["closed_trades"].cell(r, 2).value for r in range(2, export_wb["closed_trades"].max_row + 1)] == ["recent", "old"]

    full_wb = load_workbook(paths.history_path())
    assert full_wb["closed_trades"].max_row == 3


def test_export_skips_unknown_legacy_closed_trades(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path)
    now = datetime.now()
    wb = _new_history_workbook()

    closed_headers = history.SHEETS["closed_trades"]
    wb["closed_trades"].append(
        _row(
            closed_headers,
            **{
                "Recorded At": now.isoformat(timespec="seconds"),
                "Ticket": "unknown-date",
                "Symbol": "ETHUSD",
                "Session ID": "LEGACY",
                "Profit": "10",
            },
        )
    )
    wb["closed_trades"].append(
        _row(
            closed_headers,
            **{
                "Recorded At": now.isoformat(timespec="seconds"),
                "Ticket": "known-date",
                "Symbol": "ETHUSD",
                "Exit Time": now.isoformat(timespec="seconds"),
                "Session ID": "20260610_100000",
                "Profit": "10",
            },
        )
    )
    wb.save(paths.history_path())

    result = history.build_export_workbook(export_days=7)
    assert result["closed_trades"] == 1
    export_wb = load_workbook(paths.export_path())
    assert [export_wb["closed_trades"].cell(r, 2).value for r in range(2, export_wb["closed_trades"].max_row + 1)] == ["known-date"]


def test_export_workbook_compacts_config_snapshots(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path)
    now = datetime.now()
    wb = _new_history_workbook()

    closed_headers = history.SHEETS["closed_trades"]
    wb["closed_trades"].append(
        _row(
            closed_headers,
            **{
                "Recorded At": now.isoformat(timespec="seconds"),
                "Ticket": "recent",
                "Symbol": "ETHUSD",
                "Exit Time": now.isoformat(timespec="seconds"),
                "Profit": "10",
                "Config Snapshot ID": "used-snap",
            },
        )
    )
    snapshot_headers = history.SHEETS["config_snapshots"]
    for snapshot_id in ("used-snap", "unused-snap", "latest-snap"):
        wb["config_snapshots"].append(
            _row(
                snapshot_headers,
                **{
                    "Timestamp": now.isoformat(timespec="seconds"),
                    "Snapshot ID": snapshot_id,
                    "Reason": "test",
                    "Account ID": "TEST",
                    "Snapshot JSON": "X" * 30000,
                },
            )
        )
    wb.save(paths.history_path())

    result = history.build_export_workbook(export_days=7)

    assert result["ok"] is True
    export_wb = load_workbook(paths.export_path())
    exported_ids = [
        export_wb["config_snapshots"].cell(row, 2).value
        for row in range(2, export_wb["config_snapshots"].max_row + 1)
    ]
    assert exported_ids == ["used-snap", "latest-snap"]
    for row in range(2, export_wb["config_snapshots"].max_row + 1):
        text = export_wb["config_snapshots"].cell(row, 5).value
        assert len(text) < 7000
        assert "truncated_for_advisor_export" in text


def test_master_csv_sync_uses_real_session_date_for_old_rows(monkeypatch, tmp_path):
    import csv
    import core.storage_manager as storage_manager

    monkeypatch.chdir(tmp_path)
    storage_manager.set_active_account("TEST_ACCOUNT")
    with open(storage_manager.MASTER_LOG_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Time", "Ticket", "Symbol", "Type", "Vol", "Entry", "SL", "TP", "Fee", "PnL ($)", "Reason", "Market Mode", "Trigger", "Session_ID", "MAE ($)", "MFE ($)"])
        writer.writerow(["08:34:49 -> 11:12:38", "csv-1", "ETHUSD", "BUY", "1", "1", "1", "1", "0", "10", "Manual_Close", "ANY", "[USER]", "20260609_111238", "0", "10"])

    synced = history.sync_from_master_csv()
    wb = load_workbook(paths.history_path())
    ws = wb["closed_trades"]

    assert synced == 1
    assert ws.cell(2, 6).value == "2026-06-09T08:34:49"
    assert ws.cell(2, 7).value == "2026-06-09T11:12:38"


def test_api_client_reads_advisor_export_not_full_history(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path)
    full_wb = _new_history_workbook()
    export_wb = _new_history_workbook()
    headers = history.SHEETS["closed_trades"]
    full_wb["closed_trades"].append(_row(headers, **{"Ticket": "full-only"}))
    export_wb["closed_trades"].append(_row(headers, **{"Ticket": "export-only"}))
    full_wb.save(paths.history_path())
    export_wb.save(paths.export_path())

    text = api_client._workbook_text(limit_rows=10)
    assert "export-only" in text
    assert "full-only" not in text


def test_api_client_saves_latest_response_and_history_snapshot(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path)
    export_wb = _new_history_workbook()
    export_wb.save(paths.export_path())
    paths.user_context_path().replace("\\", "/")
    with open(paths.technical_settings_path(), "w", encoding="utf-8") as f:
        f.write("{}")
    with open(paths.user_context_path(), "w", encoding="utf-8") as f:
        f.write("context")

    class FakeResponse:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def read(self):
            return b'{"output_text":"advisor answer"}'

    monkeypatch.setenv("OPENAI_API_KEY", "test-key")
    monkeypatch.setattr(api_client.urllib.request, "urlopen", lambda *_args, **_kwargs: FakeResponse())

    result = api_client.send_package_to_api()

    assert result["ok"] is True
    assert result["response"].replace("\\", "/").endswith("advisor/advisor_response.md")
    assert result["response_history"].replace("\\", "/").endswith(".md")
    assert "/history/advisor_response_" in result["response_history"].replace("\\", "/")
    assert "advisor_responses" not in result["response_history"].replace("\\", "/")
    with open(paths.advisor_response_path(), "r", encoding="utf-8") as f:
        assert f.read() == "advisor answer"
    with open(result["response_history"], "r", encoding="utf-8") as f:
        assert f.read() == "advisor answer"


def test_generate_package_creates_advisor_response_template(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path)

    result = exporter.generate_advisor_package()

    assert result["ok"] is True
    with open(paths.advisor_flow_path(), "r", encoding="utf-8") as f:
        flow = f.read()
    assert "RAT-CKVN AI Advisor Flow" in flow
    with open(paths.advisor_response_path(), "r", encoding="utf-8") as f:
        text = f.read()
    assert "No API response has been saved yet." in text
    assert "AI Advisor cho RAT-CKVN" in api_client.load_advisor_prompt()
    assert api_client.load_api_settings()["technical_settings_limit"] == 1000000


def test_generate_package_clones_editable_files_from_global_templates(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path / "account")
    template_root = tmp_path / "templates"
    template_root.mkdir(parents=True)
    monkeypatch.setattr(paths, "template_root", lambda: str(template_root))
    for name in ("advisor_prompt.md", "advisor_flow.md", "user_context.md", "advisor_response.md"):
        with open(template_root / name, "w", encoding="utf-8") as f:
            f.write(f"template::{name}")

    result = exporter.generate_advisor_package()

    assert result["ok"] is True
    for path_func, name in (
        (paths.advisor_prompt_path, "advisor_prompt.md"),
        (paths.advisor_flow_path, "advisor_flow.md"),
        (paths.user_context_path, "user_context.md"),
        (paths.advisor_response_path, "advisor_response.md"),
    ):
        with open(path_func(), "r", encoding="utf-8") as f:
            assert f.read() == f"template::{name}"


def test_generate_package_does_not_overwrite_existing_editable_files(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path / "account")
    template_root = tmp_path / "templates"
    template_root.mkdir(parents=True)
    monkeypatch.setattr(paths, "template_root", lambda: str(template_root))
    for name in ("advisor_prompt.md", "advisor_flow.md", "user_context.md", "advisor_response.md"):
        with open(template_root / name, "w", encoding="utf-8") as f:
            f.write(f"template::{name}")
    paths.ensure_advisor_dirs()
    with open(paths.user_context_path(), "w", encoding="utf-8") as f:
        f.write("custom context")

    result = exporter.generate_advisor_package()

    assert result["ok"] is True
    with open(paths.user_context_path(), "r", encoding="utf-8") as f:
        assert f.read() == "custom context"


def test_api_settings_migrates_from_legacy_advisor_folder(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path)
    paths.ensure_advisor_dirs()
    with open(paths.legacy_advisor_api_settings_path(), "w", encoding="utf-8") as f:
        f.write('{"technical_settings_limit": 1234, "workbook_limit_rows": 9, "previous_response_limit": 77}')

    settings = api_client.load_api_settings()

    assert settings["technical_settings_limit"] == 1234
    assert settings["workbook_limit_rows"] == 9
    assert settings["previous_response_limit"] == 77
    assert os.path.exists(paths.advisor_api_settings_path())
    assert not os.path.exists(paths.legacy_advisor_api_settings_path())


def test_api_client_only_sends_advisor_response_when_enabled(monkeypatch, tmp_path):
    import json

    _patch_account_dir(monkeypatch, tmp_path)
    export_wb = _new_history_workbook()
    export_wb.save(paths.export_path())
    with open(paths.technical_settings_path(), "w", encoding="utf-8") as f:
        f.write("{}")
    with open(paths.user_context_path(), "w", encoding="utf-8") as f:
        f.write("context")
    with open(paths.advisor_flow_path(), "w", encoding="utf-8") as f:
        f.write("flow marker")
    with open(paths.advisor_response_path(), "w", encoding="utf-8") as f:
        f.write("previous advice marker")

    seen_inputs = []

    class FakeResponse:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def read(self):
            return b'{"output_text":"advisor answer"}'

    def fake_urlopen(req, **_kwargs):
        payload = json.loads(req.data.decode("utf-8"))
        seen_inputs.append(payload["input"])
        return FakeResponse()

    monkeypatch.setenv("OPENAI_API_KEY", "test-key")
    monkeypatch.setattr(api_client.urllib.request, "urlopen", fake_urlopen)

    assert api_client.send_package_to_api()["ok"] is True
    assert seen_inputs[-1].startswith("# advisor_flow.md")
    assert "flow marker" in seen_inputs[-1]
    assert "previous_advisor_response.md" not in seen_inputs[-1]
    assert "previous advice marker" not in seen_inputs[-1]

    with open(paths.advisor_response_path(), "w", encoding="utf-8") as f:
        f.write("previous advice marker")
    assert api_client.send_package_to_api(include_previous_response=True)["ok"] is True
    assert "previous_advisor_response.md" in seen_inputs[-1]
    assert "previous advice marker" in seen_inputs[-1]


def test_api_client_sends_max_output_tokens(monkeypatch, tmp_path):
    import json

    _patch_account_dir(monkeypatch, tmp_path)
    export_wb = _new_history_workbook()
    export_wb.save(paths.export_path())
    with open(paths.technical_settings_path(), "w", encoding="utf-8") as f:
        f.write("{}")
    with open(paths.user_context_path(), "w", encoding="utf-8") as f:
        f.write("context")
    with open(paths.advisor_flow_path(), "w", encoding="utf-8") as f:
        f.write("flow marker")
    api_client.save_api_settings({"max_output_tokens": 4096})

    seen_payloads = []

    class FakeResponse:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def read(self):
            return b'{"output_text":"advisor answer"}'

    def fake_urlopen(req, **_kwargs):
        seen_payloads.append(json.loads(req.data.decode("utf-8")))
        return FakeResponse()

    monkeypatch.setenv("OPENAI_API_KEY", "test-key")
    monkeypatch.setattr(api_client.urllib.request, "urlopen", fake_urlopen)

    assert api_client.send_package_to_api()["ok"] is True
    assert seen_payloads[-1]["max_output_tokens"] == 4096


def test_api_client_local_tpm_guard_blocks_second_large_request(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path)
    export_wb = _new_history_workbook()
    export_wb.save(paths.export_path())
    with open(paths.technical_settings_path(), "w", encoding="utf-8") as f:
        f.write("{}")
    with open(paths.user_context_path(), "w", encoding="utf-8") as f:
        f.write("context")
    with open(paths.advisor_flow_path(), "w", encoding="utf-8") as f:
        f.write("flow marker")

    class FakeResponse:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def read(self):
            return b'{"output_text":"advisor answer"}'

    calls = {"count": 0}

    def fake_urlopen(*_args, **_kwargs):
        calls["count"] += 1
        return FakeResponse()

    monkeypatch.setenv("OPENAI_API_KEY", "test-key")
    monkeypatch.setattr(api_client.urllib.request, "urlopen", fake_urlopen)
    monkeypatch.setattr(api_client, "MODEL_TPM_LIMITS", {"gpt-5.4-mini": 15000})
    monkeypatch.setattr(api_client, "LOCAL_REQUEST_OVERHEAD_TOKENS", 0)

    first = api_client.send_package_to_api()
    second = api_client.send_package_to_api()

    assert first["ok"] is True
    assert second["ok"] is False
    assert "local TPM guard" in second["error"]
    assert calls["count"] == 1


def test_api_client_sends_web_search_tool_by_default(monkeypatch, tmp_path):
    import json

    _patch_account_dir(monkeypatch, tmp_path)
    export_wb = _new_history_workbook()
    export_wb.save(paths.export_path())
    with open(paths.technical_settings_path(), "w", encoding="utf-8") as f:
        f.write("{}")
    with open(paths.user_context_path(), "w", encoding="utf-8") as f:
        f.write("context")
    with open(paths.advisor_flow_path(), "w", encoding="utf-8") as f:
        f.write("flow marker")

    seen_payloads = []

    class FakeResponse:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def read(self):
            return b'{"output_text":"advisor answer"}'

    def fake_urlopen(req, **_kwargs):
        seen_payloads.append(json.loads(req.data.decode("utf-8")))
        return FakeResponse()

    monkeypatch.setenv("OPENAI_API_KEY", "test-key")
    monkeypatch.setattr(api_client.urllib.request, "urlopen", fake_urlopen)

    assert api_client.send_package_to_api()["ok"] is True
    assert seen_payloads[-1]["tools"] == [{"type": "web_search"}]


def test_api_client_omits_web_search_tool_when_disabled(monkeypatch, tmp_path):
    import json

    _patch_account_dir(monkeypatch, tmp_path)
    export_wb = _new_history_workbook()
    export_wb.save(paths.export_path())
    with open(paths.technical_settings_path(), "w", encoding="utf-8") as f:
        f.write("{}")
    with open(paths.user_context_path(), "w", encoding="utf-8") as f:
        f.write("context")
    with open(paths.advisor_flow_path(), "w", encoding="utf-8") as f:
        f.write("flow marker")
    api_client.save_api_settings({"web_search_enabled": False})

    seen_payloads = []

    class FakeResponse:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def read(self):
            return b'{"output_text":"advisor answer"}'

    def fake_urlopen(req, **_kwargs):
        seen_payloads.append(json.loads(req.data.decode("utf-8")))
        return FakeResponse()

    monkeypatch.setenv("OPENAI_API_KEY", "test-key")
    monkeypatch.setattr(api_client.urllib.request, "urlopen", fake_urlopen)

    assert api_client.send_package_to_api()["ok"] is True
    assert "tools" not in seen_payloads[-1]
    assert api_client.load_api_settings()["web_search_enabled"] is False


def test_api_client_blocks_payload_that_exceeds_context(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path)
    export_wb = _new_history_workbook()
    export_wb.save(paths.export_path())
    with open(paths.technical_settings_path(), "w", encoding="utf-8") as f:
        f.write("{}")
    with open(paths.user_context_path(), "w", encoding="utf-8") as f:
        f.write("context")
    with open(paths.advisor_flow_path(), "w", encoding="utf-8") as f:
        f.write("flow marker")

    monkeypatch.setenv("OPENAI_API_KEY", "test-key")
    # Ép context tí hon qua catalog provider (cơ chế mới thay cho MODEL_CONTEXT_TOKENS).
    import copy
    tiny = copy.deepcopy(api_client._providers())
    tiny["openai"]["context_tokens"] = {"gpt-5.4-mini": 10}
    monkeypatch.setattr(api_client.config, "AI_ADVISOR_PROVIDERS", tiny)

    def fail_urlopen(*_args, **_kwargs):
        raise AssertionError("urlopen should not be called when payload is too large")

    monkeypatch.setattr(api_client.urllib.request, "urlopen", fail_urlopen)

    result = api_client.send_package_to_api()

    assert result["ok"] is False
    assert "payload too large" in result["error"]
    assert result["estimate"]["fits_context"] is False


def test_api_client_uses_large_limit_for_technical_settings(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path)
    export_wb = _new_history_workbook()
    export_wb.save(paths.export_path())
    with open(paths.advisor_flow_path(), "w", encoding="utf-8") as f:
        f.write("flow")
    with open(paths.technical_settings_path(), "w", encoding="utf-8") as f:
        f.write("A" * 150000)
    with open(paths.user_context_path(), "w", encoding="utf-8") as f:
        f.write("context")

    text = api_client.build_api_input()

    assert "A" * 130000 in text


def test_api_client_estimates_payload(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path)
    export_wb = _new_history_workbook()
    export_wb.save(paths.export_path())
    with open(paths.advisor_flow_path(), "w", encoding="utf-8") as f:
        f.write("flow")
    with open(paths.technical_settings_path(), "w", encoding="utf-8") as f:
        f.write("{}")
    with open(paths.user_context_path(), "w", encoding="utf-8") as f:
        f.write("context")

    estimate = api_client.estimate_api_payload()

    assert estimate["tokens"] > 0
    assert estimate["input_cost_usd"] > 0
    assert estimate["model"] == "gpt-5.4-mini"
    assert estimate["web_search_enabled"] is True
    names = [item["name"] for item in estimate["breakdown"]]
    assert "advisor_prompt.md" in names
    assert "advisor_flow.md" in names
    assert "technical_settings.json" in names
    assert "advisor_export.xlsx" in names
    assert "user_context.md" in names
    assert "advisor_response.md" not in names


def test_api_client_reads_prompt_and_limits_from_advisor_files(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path)
    export_wb = _new_history_workbook()
    export_wb.save(paths.export_path())
    with open(paths.advisor_flow_path(), "w", encoding="utf-8") as f:
        f.write("flow")
    with open(paths.technical_settings_path(), "w", encoding="utf-8") as f:
        f.write("B" * 2000)
    with open(paths.user_context_path(), "w", encoding="utf-8") as f:
        f.write("context")
    api_client.save_api_settings(
        {
            "model": "not-a-real-model",
            "advisor_prompt_limit": 1000,
            "advisor_flow_limit": 1000,
            "user_context_limit": 1000,
            "technical_settings_limit": 1200,
            "workbook_limit_rows": 3,
            "previous_response_limit": 7,
        }
    )
    api_client.save_advisor_prompt("custom opening prompt")
    with open(paths.advisor_response_path(), "w", encoding="utf-8") as f:
        f.write("response marker")

    text = api_client.build_api_input(include_previous_response=True)

    assert "B" * 1200 in text
    assert "B" * 1300 not in text
    assert "response" in text
    assert "response marker" not in text
    assert api_client.load_advisor_prompt() == "custom opening prompt"
    estimate = api_client.estimate_api_payload(include_previous_response=True)
    assert "advisor_response.md" in [item["name"] for item in estimate["breakdown"]]
    assert estimate["model"] == "gpt-5.4-mini"


def test_api_client_applies_prompt_flow_and_context_limits(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path)
    export_wb = _new_history_workbook()
    export_wb.save(paths.export_path())
    api_client.save_api_settings(
        {
            "advisor_prompt_limit": 1200,
            "advisor_flow_limit": 1300,
            "user_context_limit": 1400,
            "technical_settings_limit": 1500,
        }
    )
    with open(paths.advisor_prompt_path(), "w", encoding="utf-8") as f:
        f.write("P" * 2000)
    with open(paths.advisor_flow_path(), "w", encoding="utf-8") as f:
        f.write("F" * 2000)
    with open(paths.user_context_path(), "w", encoding="utf-8") as f:
        f.write("U" * 2000)
    with open(paths.technical_settings_path(), "w", encoding="utf-8") as f:
        f.write("T" * 2000)

    assert len(api_client.load_advisor_prompt()) == 1200
    text = api_client.build_api_input()

    assert "F" * 1300 in text
    assert "F" * 1400 not in text
    assert "U" * 1400 in text
    assert "U" * 1500 not in text
    assert "T" * 1500 in text
    assert "T" * 1600 not in text


def test_technical_snapshot_limits_active_by_symbol_to_relevant_symbols(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path)
    live_path = tmp_path / "live_signals.json"
    with open(live_path, "w", encoding="utf-8") as f:
        f.write('{"pending_signals":[{"symbol":"SSI","action":"BUY","signal_class":"ENTRY"}]}')

    import core.storage_manager as storage_manager

    monkeypatch.setattr(config_snapshot.config, "CKCS_WATCHLIST", ["FPT", "SSI", "VCB"], raising=False)
    monkeypatch.setattr(storage_manager, "load_brain_settings", lambda: {"BOT_ACTIVE_SYMBOLS": ["FPT"]})
    monkeypatch.setattr(
        storage_manager,
        "get_brain_settings_for_symbol",
        lambda symbol: {"symbol_marker": symbol},
    )
    monkeypatch.setattr(
        config_snapshot,
        "_source_paths",
        lambda: {
            "live_signals": str(live_path),
        },
    )

    snapshot = config_snapshot.build_snapshot(reason="test")
    settings = snapshot["settings"]

    assert settings["active_by_symbol"] == {
        "FPT": {"symbol_marker": "FPT"},
        "SSI": {"symbol_marker": "SSI"},
    }
    assert "VCB" in settings["omitted_symbols"]
    assert settings["relevant_symbols"] == ["FPT", "SSI"]


def test_technical_snapshot_includes_symbols_from_advisor_workbook(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path)
    wb = _new_history_workbook()
    wb["open_trades"].append(["", "", "SSI"])
    wb.save(paths.export_path())

    import core.storage_manager as storage_manager

    monkeypatch.setattr(config_snapshot.config, "CKCS_WATCHLIST", ["FPT", "SSI", "VCB"], raising=False)
    monkeypatch.setattr(storage_manager, "load_brain_settings", lambda: {"BOT_ACTIVE_SYMBOLS": ["FPT"]})
    monkeypatch.setattr(
        storage_manager,
        "get_brain_settings_for_symbol",
        lambda symbol: {"symbol_marker": symbol},
    )
    monkeypatch.setattr(config_snapshot, "_source_paths", lambda: {})

    settings = config_snapshot.build_snapshot(reason="test")["settings"]

    assert set(settings["active_by_symbol"]) == {"FPT", "SSI"}
    assert settings["omitted_symbols"] == ["VN30F1M", "VCB"]


# --- AI Advisor đa provider/model (OpenAI + Claude) ---

def test_advisor_provider_catalog_has_openai_and_anthropic():
    providers = api_client._providers()
    assert "openai" in providers and "anthropic" in providers
    assert api_client.default_provider() == "openai"
    assert "gpt-5.4-mini" in api_client.models_for("openai")
    assert "claude-sonnet-4-6" in api_client.models_for("anthropic")


def test_advisor_normalize_model_per_provider():
    assert api_client.normalize_model("gpt-5.4", "openai") == "gpt-5.4"
    # model lạ -> rơi về default model của provider đó
    assert api_client.normalize_model("khong-co", "anthropic") == "claude-sonnet-4-6"
    assert api_client.normalize_model("claude-opus-4-8", "anthropic") == "claude-opus-4-8"
    assert api_client.normalize_provider("xxx") == "openai"


def test_advisor_build_request_openai_shape_unchanged():
    endpoint, headers, payload = api_client._build_request(
        "openai", "gpt-5.4-mini", "SYS", "BODY", 8000, True, "KEY"
    )
    assert endpoint.endswith("/v1/responses")
    assert headers["Authorization"] == "Bearer KEY"
    assert payload["instructions"] == "SYS" and payload["input"] == "BODY"
    assert payload["max_output_tokens"] == 8000
    assert payload["tools"] == [{"type": "web_search"}]


def test_advisor_build_request_anthropic_shape():
    endpoint, headers, payload = api_client._build_request(
        "anthropic", "claude-sonnet-4-6", "SYS", "BODY", 4096, True, "KEY"
    )
    assert endpoint.endswith("/v1/messages")
    assert headers["x-api-key"] == "KEY"
    assert headers["anthropic-version"]
    assert payload["system"] == "SYS"
    assert payload["messages"] == [{"role": "user", "content": "BODY"}]
    assert payload["max_tokens"] == 4096
    assert payload["tools"]  # web_search tool có mặt


def test_advisor_parse_response_per_provider():
    assert api_client._parse_response("anthropic", {"content": [{"type": "text", "text": "hi"}]}) == "hi"
    assert api_client._parse_response("openai", {"output_text": "ok"}) == "ok"


def test_advisor_settings_roundtrip_provider(monkeypatch, tmp_path):
    _patch_account_dir(monkeypatch, tmp_path)
    saved = api_client.save_api_settings({"provider": "anthropic", "model": "claude-opus-4-8"})
    assert saved["provider"] == "anthropic"
    assert saved["model"] == "claude-opus-4-8"
    loaded = api_client.load_api_settings()
    assert loaded["provider"] == "anthropic"
    assert loaded["model"] == "claude-opus-4-8"
    # mặc định (chưa lưu gì) -> provider openai
    api_client.save_api_settings({})
    assert api_client.load_api_settings()["provider"] == "openai"
