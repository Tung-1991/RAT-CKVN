# -*- coding: utf-8 -*-
import json
import hashlib
import logging
import os
import random
import re
import shutil
import time
import urllib.error
import urllib.request

import config

from . import history, paths

logger = logging.getLogger("RAT_CKVN")


DEFAULT_PROMPT = """Bạn là AI Advisor cho RAT-CKVN. Luôn trả lời bằng tiếng Việt, chuyên nghiệp, sắc gọn và dựa trên bằng chứng.
Đọc advisor_flow.md trước để hiểu RAT-CKVN, sau đó đọc user_context.md, technical_settings.json, advisor_export.xlsx và previous_advisor_response.md nếu có.
Nếu web_search được bật, bắt buộc kiểm tra bối cảnh thị trường mới cho symbol active hoặc symbol có trade trong export; chỉ giữ thông tin web có tác động trực tiếp tới chẩn đoán RAT-CKVN.
Tách rõ dữ liệu nội bộ RAT-CKVN với bối cảnh thị trường/web. Không viết bản tin tổng hợp, không lặp lại quá nhiều số liệu nếu đã nêu ở evidence.
Không dùng markdown bold/italic, không dùng ký tự **, không paste URL dài trong thân bài, không dùng bảng Markdown. Ưu tiên report khoảng 700-1000 từ và 1-2 Telegram chunks; nếu dữ liệu phức tạp, được dài hơn nhưng phải gọn và không lặp số liệu.
Không đề xuất đặt lệnh tự động, không yêu cầu bot tự sửa config, và không biến lời khuyên trước đó thành sự thật nếu chưa kiểm chứng bằng dữ liệu hiện tại."""

TECHNICAL_SETTINGS_LIMIT = 1000000
PROMPT_LIMIT = 200000
ADVISOR_FLOW_LIMIT = 200000
USER_CONTEXT_LIMIT = 100000
PREVIOUS_RESPONSE_LIMIT = 60000
WORKBOOK_LIMIT_ROWS = 80
DEFAULT_MAX_OUTPUT_TOKENS = 8000
WORKBOOK_CELL_LIMIT = 6000
# Fallback nếu config.py thiếu catalog (giữ api_client tự chạy được).
_FALLBACK_PROVIDERS = {
    "openai": {
        "label": "OpenAI",
        "endpoint": "https://api.openai.com/v1/responses",
        "env_key": "OPENAI_API_KEY",
        "models": ["gpt-5.6-terra", "gpt-5.6", "gpt-5.6-sol", "gpt-5.6-luna", "gpt-5.4-mini", "gpt-5.4", "gpt-5.5"],
        "default_model": "gpt-5.6",
        "context_tokens": {
            "gpt-5.6-terra": 1050000,
            "gpt-5.6": 1050000,
            "gpt-5.6-sol": 1050000,
            "gpt-5.6-luna": 1050000,
            "gpt-5.4-mini": 400000,
            "gpt-5.4": 1000000,
            "gpt-5.5": 1000000,
        },
        "pricing": {
            "gpt-5.6-terra": {"input": 2.50, "output": 15.00},
            "gpt-5.6": {"input": 5.00, "output": 30.00},
            "gpt-5.6-sol": {"input": 5.00, "output": 30.00},
            "gpt-5.6-luna": {"input": 1.00, "output": 6.00},
            "gpt-5.4-mini": {"input": 0.75, "output": 4.50},
            "gpt-5.4": {"input": 2.50, "output": 15.00},
            "gpt-5.5": {"input": 5.00, "output": 30.00},
        },
    },
    "anthropic": {
        "label": "Claude (Anthropic)",
        "endpoint": "https://api.anthropic.com/v1/messages",
        "env_key": "ANTHROPIC_API_KEY",
        "models": ["claude-opus-4-8", "claude-sonnet-4-6", "claude-haiku-4-5"],
        "default_model": "claude-sonnet-4-6",
        "context_tokens": {"claude-opus-4-8": 200000, "claude-sonnet-4-6": 200000, "claude-haiku-4-5": 200000},
        "pricing": {
            "claude-opus-4-8": {"input": 5.00, "output": 25.00},
            "claude-sonnet-4-6": {"input": 3.00, "output": 15.00},
            "claude-haiku-4-5": {"input": 1.00, "output": 5.00},
        },
    },
}


def _providers():
    cfg = getattr(config, "AI_ADVISOR_PROVIDERS", None)
    return cfg if isinstance(cfg, dict) and cfg else _FALLBACK_PROVIDERS


def default_provider():
    p = str(getattr(config, "AI_ADVISOR_DEFAULT_PROVIDER", "openai") or "openai").strip().lower()
    return p if p in _providers() else next(iter(_providers()))


def normalize_provider(value):
    p = str(value or "").strip().lower()
    return p if p in _providers() else default_provider()


def provider_config(provider=None):
    return _providers()[normalize_provider(provider)]


def models_for(provider=None):
    return list(provider_config(provider).get("models", []))


def provider_labels():
    return {p: cfg.get("label", p) for p, cfg in _providers().items()}


DEFAULT_PROVIDER = default_provider()
DEFAULT_MODEL = provider_config(DEFAULT_PROVIDER).get("default_model", "gpt-5.6")
# Backward-compat: model của provider mặc định.
SUPPORTED_MODELS = models_for(DEFAULT_PROVIDER)
LOCAL_TPM_WINDOW_SECONDS = 60
LOCAL_REQUEST_OVERHEAD_TOKENS = 10000
# Không đoán tier TPM của tài khoản OpenAI. Mặc định để trống và xử lý 429/
# Retry-After từ API; biến này vẫn giữ để deployment có thể cấu hình guard cục bộ.
MODEL_TPM_LIMITS = {}


DEFAULT_API_SETTINGS = {
    "settings_version": 2,
    "provider": DEFAULT_PROVIDER,
    "model": DEFAULT_MODEL,
    "reasoning_effort": "medium",
    "web_search_enabled": True,
    "advisor_prompt_limit": PROMPT_LIMIT,
    "advisor_flow_limit": ADVISOR_FLOW_LIMIT,
    "user_context_limit": USER_CONTEXT_LIMIT,
    "technical_settings_limit": TECHNICAL_SETTINGS_LIMIT,
    "previous_response_limit": PREVIOUS_RESPONSE_LIMIT,
    "workbook_limit_rows": WORKBOOK_LIMIT_ROWS,
    "max_output_tokens": DEFAULT_MAX_OUTPUT_TOKENS,
}


def _safe_int(value, default, min_value=1, max_value=5000000):
    try:
        parsed = int(float(value))
    except Exception:
        return default
    return max(min_value, min(max_value, parsed))


def _safe_bool(value, default=False):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"1", "true", "yes", "on", "y"}:
            return True
        if normalized in {"0", "false", "no", "off", "n"}:
            return False
    return bool(default)


def normalize_model(value, provider=None):
    models = models_for(provider)
    model = str(value or "").strip()
    if model in models:
        return model
    return provider_config(provider).get("default_model") or (models[0] if models else DEFAULT_MODEL)


def normalize_reasoning_effort(value):
    effort = str(value or "medium").strip().lower()
    return effort if effort in {"none", "low", "medium", "high", "xhigh", "max"} else "medium"


def _build_request(
    provider,
    model,
    prompt,
    body_text,
    max_output_tokens,
    web_search,
    api_key,
    reasoning_effort="medium",
):
    """Trả (endpoint, headers, payload) theo từng provider. Hàm thuần để test được."""
    pcfg = provider_config(provider)
    provider = normalize_provider(provider)
    endpoint = _get_env_value("ADVISOR_API_URL") or pcfg["endpoint"]
    if provider == "anthropic":
        payload = {
            "model": model,
            "max_tokens": int(max_output_tokens),
            "system": prompt,
            "messages": [{"role": "user", "content": body_text}],
        }
        if web_search:
            payload["tools"] = [{"type": "web_search_20250305", "name": "web_search"}]
        headers = {
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        }
    else:  # openai (mặc định)
        payload = {
            "model": model,
            "instructions": prompt,
            "input": body_text,
            "max_output_tokens": int(max_output_tokens),
        }
        if str(model).startswith("gpt-5.6"):
            payload["reasoning"] = {"effort": normalize_reasoning_effort(reasoning_effort)}
        if web_search:
            payload["tools"] = [{"type": "web_search"}]
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        }
    return endpoint, headers, payload


def _parse_response(provider, data):
    """Trích text trả lời theo từng provider. Hàm thuần để test được."""
    if normalize_provider(provider) == "anthropic":
        parts = []
        for block in data.get("content", []) or []:
            if isinstance(block, dict) and block.get("type") == "text":
                parts.append(block.get("text", ""))
        return "\n".join(parts).strip()
    # openai
    text = data.get("output_text")
    if not text:
        parts = []
        for item in data.get("output", []) or []:
            for content in item.get("content", []) or []:
                if content.get("type") == "output_text":
                    parts.append(content.get("text", ""))
        text = "\n".join(parts).strip()
    return text or ""


def _stdout_log(message):
    try:
        logger.info("[AI ADVISOR API] %s", message)
    except Exception:
        pass


def _get_env_value(name):
    name = str(name or "").strip()
    if not name:
        return ""
    value = os.environ.get(name, "")
    if value:
        return value
    if os.name != "nt":
        return ""
    try:
        import winreg

        locations = [
            (winreg.HKEY_CURRENT_USER, "Environment"),
            (
                winreg.HKEY_LOCAL_MACHINE,
                r"SYSTEM\CurrentControlSet\Control\Session Manager\Environment",
            ),
        ]
        for root, path in locations:
            try:
                with winreg.OpenKey(root, path) as key:
                    registry_value, _value_type = winreg.QueryValueEx(key, name)
                if registry_value:
                    return str(registry_value)
            except OSError:
                continue
    except Exception:
        pass
    return ""


def _urlopen(req, timeout=None):
    """Open HTTPS with Python's default certificate verification enabled."""
    timeout = float(
        timeout
        if timeout is not None
        else getattr(config, "ADVISOR_API_TIMEOUT_SECONDS", 300.0)
    )
    return urllib.request.urlopen(req, timeout=timeout)


def _extract_citations(data):
    citations = []
    seen = set()
    for item in (data or {}).get("output", []) or []:
        for content in item.get("content", []) if isinstance(item, dict) else []:
            for annotation in content.get("annotations", []) if isinstance(content, dict) else []:
                if not isinstance(annotation, dict):
                    continue
                citation = annotation.get("url_citation") if isinstance(annotation.get("url_citation"), dict) else annotation
                url = str(citation.get("url") or "").strip()
                if not url or url in seen:
                    continue
                seen.add(url)
                citations.append({"title": str(citation.get("title") or "Nguồn").strip(), "url": url})
    return citations


def _append_citations(text, citations):
    if not citations:
        return text
    lines = [text.rstrip(), "", "Nguồn tham khảo:"]
    for item in citations[:20]:
        lines.append(f"- [{item['title']}]({item['url']})")
    return "\n".join(lines).strip()


def _safe_error_text(value, api_key=""):
    text = str(value or "")
    if api_key:
        text = text.replace(str(api_key), "[REDACTED]")
    text = re.sub(r"(?i)(bearer\s+)[A-Za-z0-9._~+\-/=]+", r"\1[REDACTED]", text)
    text = re.sub(r"(?i)(api[_-]?key|token|secret|password|otp)(\s*[:=]\s*)[^\s,;]+", r"\1\2[REDACTED]", text)
    return text[:4000]


def _sanitize_external_text(value):
    text = str(value or "")
    for name in (
        "OPENAI_API_KEY",
        "ANTHROPIC_API_KEY",
        "DNSE_API_KEY",
        "DNSE_API_SECRET",
        "DNSE_CUSTODY_CODE",
        "DNSE_ACCOUNT_NO",
        "DNSE_STOCK_ACCOUNT_NO",
        "DNSE_DERIVATIVE_ACCOUNT_NO",
    ):
        secret = str(os.getenv(name, "") or getattr(config, name, "") or "").strip()
        if not secret:
            continue
        replacement = "[REDACTED]"
        if "ACCOUNT_NO" in name or name == "DNSE_CUSTODY_CODE":
            replacement = "ACCOUNT#" + hashlib.sha256(secret.encode("utf-8")).hexdigest()[:10]
        text = text.replace(secret, replacement)
    text = re.sub(r"(?i)(api[_-]?key|secret|trading[_-]?token|password|passcode|otp)(\s*[:=]\s*)[^\s,;]+", r"\1\2[REDACTED]", text)
    text = re.sub(r"(?<![\w])(?:[A-Za-z]:[\\/][^\r\n|]+)", "<LOCAL_PATH>", text)
    return text


def _retry_after_seconds(headers, attempt):
    try:
        value = float((headers or {}).get("Retry-After", "") or 0.0)
    except (TypeError, ValueError):
        value = 0.0
    base = value if value > 0 else min(30.0, 2.0 ** max(1, attempt))
    return base + random.uniform(0.05, min(0.5, max(0.05, base * 0.1)))


def _api_usage_path():
    return os.path.join(paths.account_dir(), "advisor_api_usage.json")


def _load_api_usage(now=None):
    now = now or time.time()
    path = _api_usage_path()
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        rows = data if isinstance(data, list) else []
    except Exception:
        return []
    return [
        row
        for row in rows
        if isinstance(row, dict)
        and now - float(row.get("ts", 0.0) or 0.0) < LOCAL_TPM_WINDOW_SECONDS
    ]


def _save_api_usage(rows):
    path = _api_usage_path()
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        tmp_path = f"{path}.tmp"
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(rows, f, indent=2, ensure_ascii=False)
        os.replace(tmp_path, path)
    except Exception:
        pass


def _requested_tokens_for_guard(estimate):
    return int(estimate.get("tokens", 0) or 0) + int(estimate.get("max_output_tokens", 0) or 0) + LOCAL_REQUEST_OVERHEAD_TOKENS


def _check_local_tpm_guard(model, requested_tokens):
    limit = MODEL_TPM_LIMITS.get(model)
    if not limit:
        return {"ok": True, "used_tokens": 0, "limit": None, "wait_seconds": 0}

    now = time.time()
    rows = _load_api_usage(now=now)
    used_tokens = sum(
        int(row.get("requested_tokens", 0) or 0)
        for row in rows
        if row.get("model") == model
    )
    if used_tokens + requested_tokens <= limit:
        return {"ok": True, "used_tokens": used_tokens, "limit": limit, "wait_seconds": 0}

    oldest = min(
        (
            float(row.get("ts", now) or now)
            for row in rows
            if row.get("model") == model
        ),
        default=now,
    )
    wait_seconds = max(1, int(LOCAL_TPM_WINDOW_SECONDS - (now - oldest)) + 1)
    return {
        "ok": False,
        "used_tokens": used_tokens,
        "limit": limit,
        "wait_seconds": wait_seconds,
    }


def _record_api_usage(model, requested_tokens):
    now = time.time()
    rows = _load_api_usage(now=now)
    rows.append(
        {
            "ts": now,
            "model": model,
            "requested_tokens": int(requested_tokens),
        }
    )
    _save_api_usage(rows)


def load_api_settings():
    settings = dict(DEFAULT_API_SETTINGS)
    path = paths.advisor_api_settings_path()
    legacy_path = paths.legacy_advisor_api_settings_path()
    source_path = path if os.path.exists(path) else legacy_path
    loaded = {}
    if os.path.exists(source_path):
        try:
            with open(source_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                loaded = dict(data)
                settings.update(data)
        except Exception:
            pass
    settings["technical_settings_limit"] = _safe_int(
        settings.get("technical_settings_limit"),
        TECHNICAL_SETTINGS_LIMIT,
        min_value=1000,
    )
    settings["advisor_prompt_limit"] = _safe_int(
        settings.get("advisor_prompt_limit"),
        PROMPT_LIMIT,
        min_value=1000,
    )
    settings["advisor_flow_limit"] = _safe_int(
        settings.get("advisor_flow_limit"),
        ADVISOR_FLOW_LIMIT,
        min_value=1000,
    )
    settings["user_context_limit"] = _safe_int(
        settings.get("user_context_limit"),
        USER_CONTEXT_LIMIT,
        min_value=1000,
    )
    settings["previous_response_limit"] = _safe_int(
        settings.get("previous_response_limit"),
        PREVIOUS_RESPONSE_LIMIT,
        min_value=0,
    )
    settings["workbook_limit_rows"] = _safe_int(
        settings.get("workbook_limit_rows"),
        WORKBOOK_LIMIT_ROWS,
        min_value=1,
        max_value=10000,
    )
    settings["max_output_tokens"] = _safe_int(
        settings.get("max_output_tokens"),
        DEFAULT_MAX_OUTPUT_TOKENS,
        min_value=1024,
        max_value=128000,
    )
    settings["provider"] = normalize_provider(settings.get("provider"))
    loaded_version = _safe_int(loaded.get("settings_version", 1), 1, min_value=1, max_value=100)
    if (
        loaded_version < 2
        and settings["provider"] == "openai"
        and str(settings.get("model") or "") == "gpt-5.4-mini"
    ):
        settings["model"] = "gpt-5.6"
    settings["model"] = normalize_model(settings.get("model"), settings["provider"])
    settings["reasoning_effort"] = normalize_reasoning_effort(settings.get("reasoning_effort"))
    settings["settings_version"] = 2
    settings["web_search_enabled"] = _safe_bool(settings.get("web_search_enabled"), True)
    needs_save = bool(loaded) and loaded != settings
    if needs_save and source_path == path:
        try:
            save_api_settings(settings)
        except Exception:
            pass
    if os.path.exists(legacy_path) and (source_path == legacy_path or os.path.exists(path)):
        try:
            save_api_settings(settings)
            os.remove(legacy_path)
        except Exception:
            pass
    return settings


def save_api_settings(settings):
    paths.ensure_advisor_dirs()
    clean = dict(DEFAULT_API_SETTINGS)
    clean.update(settings or {})
    clean = load_api_settings_from_dict(clean)
    with open(paths.advisor_api_settings_path(), "w", encoding="utf-8") as f:
        json.dump(clean, f, indent=2, ensure_ascii=False)
    return clean


def load_api_settings_from_dict(data):
    data = data or {}
    provider = normalize_provider(data.get("provider"))
    return {
        "settings_version": 2,
        "provider": provider,
        "model": normalize_model(data.get("model"), provider),
        "reasoning_effort": normalize_reasoning_effort(data.get("reasoning_effort")),
        "web_search_enabled": _safe_bool(data.get("web_search_enabled"), True),
        "advisor_prompt_limit": _safe_int(
            data.get("advisor_prompt_limit"),
            PROMPT_LIMIT,
            min_value=1000,
        ),
        "advisor_flow_limit": _safe_int(
            data.get("advisor_flow_limit"),
            ADVISOR_FLOW_LIMIT,
            min_value=1000,
        ),
        "user_context_limit": _safe_int(
            data.get("user_context_limit"),
            USER_CONTEXT_LIMIT,
            min_value=1000,
        ),
        "technical_settings_limit": _safe_int(
            data.get("technical_settings_limit"),
            TECHNICAL_SETTINGS_LIMIT,
            min_value=1000,
        ),
        "previous_response_limit": _safe_int(
            data.get("previous_response_limit"),
            PREVIOUS_RESPONSE_LIMIT,
            min_value=0,
        ),
        "workbook_limit_rows": _safe_int(
            data.get("workbook_limit_rows"),
            WORKBOOK_LIMIT_ROWS,
            min_value=1,
            max_value=10000,
        ),
        "max_output_tokens": _safe_int(
            data.get("max_output_tokens"),
            DEFAULT_MAX_OUTPUT_TOKENS,
            min_value=1024,
            max_value=128000,
        ),
    }


def ensure_advisor_prompt():
    paths.ensure_advisor_dirs()
    path = paths.advisor_prompt_path()
    if not os.path.exists(path):
        template_path = paths.advisor_template_path("advisor_prompt.md")
        if os.path.exists(template_path):
            with open(template_path, "r", encoding="utf-8", errors="replace") as src:
                text = src.read()
        else:
            text = DEFAULT_PROMPT
        with open(path, "w", encoding="utf-8") as f:
            f.write(text)
    return path


def load_advisor_prompt():
    ensure_advisor_prompt()
    text = _read_text(paths.advisor_prompt_path(), limit=load_api_settings()["advisor_prompt_limit"]).strip()
    return text or DEFAULT_PROMPT


def save_advisor_prompt(text):
    paths.ensure_advisor_dirs()
    with open(paths.advisor_prompt_path(), "w", encoding="utf-8") as f:
        f.write((text or DEFAULT_PROMPT).strip() + "\n")
    return paths.advisor_prompt_path()


def _read_text(path, limit=120000):
    if not os.path.exists(path):
        return ""
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return f.read(limit)


def _workbook_text(limit_rows=None):
    try:
        from openpyxl import load_workbook

        if not os.path.exists(paths.export_path()):
            return ""
        if limit_rows is None:
            limit_rows = load_api_settings().get("workbook_limit_rows", WORKBOOK_LIMIT_ROWS)
        wb = load_workbook(paths.export_path(), data_only=True)
        chunks = []
        for name in wb.sheetnames:
            ws = wb[name]
            chunks.append(f"\n## {name}")
            max_row = min(ws.max_row, limit_rows)
            for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True):
                cells = []
                for value in row:
                    text = "" if value is None else str(value)
                    if len(text) > WORKBOOK_CELL_LIMIT:
                        text = text[:WORKBOOK_CELL_LIMIT] + "...[truncated_for_api]"
                    cells.append(text)
                chunks.append(" | ".join(cells))
        return "\n".join(chunks)
    except Exception as exc:
        return f"advisor_export.xlsx read warning: {exc}"


def build_api_sections(include_previous_response=False):
    settings = load_api_settings()
    sections = [
        ("advisor_flow.md", _read_text(paths.advisor_flow_path(), limit=settings["advisor_flow_limit"])),
        (
            "technical_settings.json",
            _read_text(paths.technical_settings_path(), limit=settings["technical_settings_limit"]),
        ),
        ("advisor_export.xlsx", _workbook_text(limit_rows=settings["workbook_limit_rows"])),
        ("user_context.md", _read_text(paths.user_context_path(), limit=settings["user_context_limit"])),
    ]
    expert_text = _read_text(paths.expert_context_path(), limit=settings["user_context_limit"])
    if expert_text.strip():
        sections.append(("expert_context.md", expert_text))
    if include_previous_response:
        sections.append(
            (
                "advisor_response.md",
                _read_text(paths.advisor_response_path(), limit=settings["previous_response_limit"]),
            )
        )
    return [(name, _sanitize_external_text(text)) for name, text in sections]


def build_api_input(include_previous_response=False):
    sections = []
    for name, text in build_api_sections(include_previous_response=include_previous_response):
        section_name = "previous_advisor_response.md" if name == "advisor_response.md" else name
        sections.extend([f"# {section_name}", text])
    return "\n\n".join(sections)


def validate_advisor_package():
    ensure_advisor_prompt()
    required = {
        "advisor_prompt.md": paths.advisor_prompt_path(),
        "advisor_flow.md": paths.advisor_flow_path(),
        "technical_settings.json": paths.technical_settings_path(),
        "advisor_export.xlsx": paths.export_path(),
        "user_context.md": paths.user_context_path(),
    }
    files = []
    missing = []
    for name, path in required.items():
        exists = os.path.isfile(path) and os.path.getsize(path) > 0
        files.append({"name": name, "path": path, "exists": exists, "bytes": os.path.getsize(path) if exists else 0})
        if not exists:
            missing.append(name)
    return {
        "ok": not missing,
        "missing": missing,
        "files": files,
        "privacy_ok": True,
        "redaction": "secrets/account identifiers/absolute paths sanitized before API input",
    }


def estimate_api_payload(include_previous_response=False):
    prompt_text = load_advisor_prompt()
    input_sections = build_api_sections(include_previous_response=include_previous_response)
    settings = load_api_settings()
    provider = settings.get("provider", DEFAULT_PROVIDER)
    model = settings.get("model", DEFAULT_MODEL)
    pcfg = provider_config(provider)
    pricing_map = pcfg.get("pricing", {})
    pricing = pricing_map.get(model) or next(iter(pricing_map.values()), {"input": 0.0, "output": 0.0})
    text = "\n\n".join(
        part
        for name, section_text in input_sections
        for part in (f"# {'previous_advisor_response.md' if name == 'advisor_response.md' else name}", section_text)
    )
    chars = len(text) + len(prompt_text)
    tokens = max(1, int(chars / 4))
    context_map = pcfg.get("context_tokens", {})
    context_tokens = context_map.get(model) or next(iter(context_map.values()), 200000)
    max_output_tokens = settings.get("max_output_tokens", DEFAULT_MAX_OUTPUT_TOKENS)
    context_remaining_tokens = context_tokens - tokens - max_output_tokens
    input_cost = (tokens / 1000000.0) * pricing["input"]
    output_2k_cost = (2000 / 1000000.0) * pricing["output"]
    output_4k_cost = (4000 / 1000000.0) * pricing["output"]
    output_limit_cost = (max_output_tokens / 1000000.0) * pricing["output"]
    breakdown = []
    prompt_chars = len(prompt_text)
    breakdown.append(
        {
            "name": "advisor_prompt.md",
            "chars": prompt_chars,
            "tokens": max(1, int(prompt_chars / 4)),
            "included": True,
        }
    )
    for name, section_text in input_sections:
        section_chars = len(section_text or "")
        breakdown.append(
            {
                "name": name,
                "chars": section_chars,
                "tokens": max(1, int(section_chars / 4)),
                "included": True,
            }
        )
    return {
        "chars": chars,
        "tokens": tokens,
        "input_cost_usd": input_cost,
        "estimated_output_2k_usd": output_2k_cost,
        "estimated_output_4k_usd": output_4k_cost,
        "estimated_output_limit_usd": output_limit_cost,
        "provider": provider,
        "model": model,
        "context_tokens": context_tokens,
        "max_output_tokens": max_output_tokens,
        "context_remaining_tokens": context_remaining_tokens,
        "fits_context": context_remaining_tokens >= 0,
        "web_search_enabled": bool(settings.get("web_search_enabled")),
        "settings": settings,
        "breakdown": breakdown,
    }


def send_package_to_api(prompt=None, include_previous_response=False):
    package = validate_advisor_package()
    if not package["ok"]:
        msg = "Missing/empty advisor package file(s): " + ", ".join(package["missing"]) + ". Generate Advisor Package first."
        history.record_event("advisor_api_invalid_package", msg, severity="WARN", payload={"missing": package["missing"]})
        return {"ok": False, "error": msg, "package": package}
    settings = load_api_settings()
    provider = settings.get("provider", DEFAULT_PROVIDER)
    pcfg = provider_config(provider)
    env_key_name = pcfg.get("env_key", "OPENAI_API_KEY")
    key = _get_env_value(env_key_name)
    if not key:
        msg = f"{env_key_name} is not configured; API mode skipped."
        _stdout_log(msg)
        history.record_event("advisor_api_missing_key", msg, severity="WARN")
        return {"ok": False, "error": msg}

    model = settings.get("model", DEFAULT_MODEL)
    if model not in models_for(provider):
        msg = f"Unsupported model '{model}' for provider '{provider}'. Choose one of: {', '.join(models_for(provider))}"
        _stdout_log(msg)
        history.record_event("advisor_api_bad_model", msg, severity="ERROR", payload={"model": model})
        return {"ok": False, "error": msg}
    endpoint = _get_env_value("ADVISOR_API_URL") or pcfg.get("endpoint", "https://api.openai.com/v1/responses")
    body_text = build_api_input(include_previous_response=include_previous_response)
    estimate = estimate_api_payload(include_previous_response=include_previous_response)
    if not estimate.get("fits_context"):
        msg = (
            f"Advisor API payload too large for {model}: "
            f"~{estimate.get('tokens')} input tokens + "
            f"{estimate.get('max_output_tokens')} output reserve > "
            f"{estimate.get('context_tokens')} context tokens. "
            "Lower advisor_export rows/sheet or reduce max output tokens."
        )
        _stdout_log(msg)
        history.record_event(
            "advisor_api_payload_too_large",
            msg,
            severity="ERROR",
            payload={"model": model, "estimate": estimate},
        )
        return {"ok": False, "error": msg, "estimate": estimate}
    requested_tokens = _requested_tokens_for_guard(estimate)
    guard = _check_local_tpm_guard(model, requested_tokens)
    if not guard.get("ok"):
        msg = (
            f"Advisor API local TPM guard: wait {guard.get('wait_seconds')}s before sending {model}. "
            f"local_used~{guard.get('used_tokens')} requested~{requested_tokens} limit={guard.get('limit')}."
        )
        _stdout_log(msg)
        history.record_event(
            "advisor_api_local_tpm_guard",
            msg,
            severity="WARN",
            payload={
                "model": model,
                "used_tokens": guard.get("used_tokens"),
                "requested_tokens": requested_tokens,
                "limit": guard.get("limit"),
                "wait_seconds": guard.get("wait_seconds"),
            },
        )
        return {"ok": False, "error": msg, "estimate": estimate, "wait_seconds": guard.get("wait_seconds")}
    _stdout_log(
        "sending "
        f"model={model} endpoint={endpoint} "
        f"chars={estimate.get('chars')} tokens~{estimate.get('tokens')} "
        f"include_response={bool(include_previous_response)}"
    )
    _endpoint2, headers, payload = _build_request(
        provider,
        model,
        _sanitize_external_text(prompt or load_advisor_prompt()),
        body_text,
        estimate.get("max_output_tokens", DEFAULT_MAX_OUTPUT_TOKENS),
        settings.get("web_search_enabled"),
        key,
        settings.get("reasoning_effort", "medium"),
    )
    req = urllib.request.Request(
        endpoint,
        data=json.dumps(payload).encode("utf-8"),
        headers=headers,
        method="POST",
    )
    _record_api_usage(model, requested_tokens)
    max_retries = max(0, int(getattr(config, "ADVISOR_API_RETRIES", 2) or 0))
    attempt = 0
    while True:
        try:
            with _urlopen(req) as resp:
                data = json.loads(resp.read().decode("utf-8"))
            text = _parse_response(provider, data)
            if not text:
                text = json.dumps(data, ensure_ascii=False, indent=2)
            citations = _extract_citations(data) if provider == "openai" else []
            text = _append_citations(text, citations)
            with open(paths.advisor_response_path(), "w", encoding="utf-8") as f:
                f.write(text)
            response_history = paths.advisor_response_history_path()
            os.makedirs(os.path.dirname(response_history), exist_ok=True)
            shutil.copy2(paths.advisor_response_path(), response_history)
            actual_model = str(data.get("model") or model) if isinstance(data, dict) else model
            usage = data.get("usage", {}) if isinstance(data, dict) else {}
            history.record_event(
                "advisor_api_response_saved",
                "Advisor API response saved",
                payload={
                    "model": actual_model,
                    "usage": usage,
                    "citations": len(citations),
                    "response_history": response_history,
                    "include_previous_response": bool(include_previous_response),
                },
            )
            _stdout_log(f"response saved model={actual_model} response={paths.advisor_response_path()} history={response_history}")
            return {
                "ok": True,
                "response": paths.advisor_response_path(),
                "response_history": response_history,
                "model": actual_model,
                "usage": usage,
                "citations": citations,
            }
        except urllib.error.HTTPError as exc:
            detail = ""
            try:
                detail = exc.read().decode("utf-8", errors="replace")
            except Exception:
                detail = ""
            retryable = exc.code == 429 or 500 <= int(exc.code) < 600
            if retryable and attempt < max_retries:
                attempt += 1
                wait_s = _retry_after_seconds(getattr(exc, "headers", {}), attempt)
                _stdout_log(f"HTTP {exc.code}; retry {attempt}/{max_retries} after {wait_s:.1f}s")
                time.sleep(wait_s)
                continue
            msg = _safe_error_text(f"HTTP {exc.code}: {detail or exc.reason}", key)
            _stdout_log(msg)
            history.record_event(
                "advisor_api_error",
                msg,
                severity="ERROR",
                payload={"model": model, "endpoint": endpoint, "status": exc.code, "attempts": attempt + 1},
            )
            return {"ok": False, "error": msg}
        except (urllib.error.URLError, TimeoutError) as exc:
            if attempt < max_retries:
                attempt += 1
                wait_s = _retry_after_seconds({}, attempt)
                _stdout_log(f"network error; retry {attempt}/{max_retries} after {wait_s:.1f}s")
                time.sleep(wait_s)
                continue
            msg = _safe_error_text(exc, key)
            history.record_event(
                "advisor_api_error",
                msg,
                severity="ERROR",
                payload={"model": model, "endpoint": endpoint, "attempts": attempt + 1},
            )
            return {"ok": False, "error": msg}
        except Exception as exc:
            msg = _safe_error_text(exc, key)
            _stdout_log(f"error: {msg}")
            history.record_event(
                "advisor_api_error",
                msg,
                severity="ERROR",
                payload={"model": model, "endpoint": endpoint, "attempts": attempt + 1},
            )
            return {"ok": False, "error": msg}
