# -*- coding: utf-8 -*-
import copy
import hashlib
import json
import os
from datetime import datetime

import config

from . import paths


def _json_safe(value):
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_json_safe(v) for v in value]
    return str(value)


def _read_json_file(path):
    if not path or not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as exc:
        return {"_advisor_read_error": str(exc)}


def _public_config_values():
    values = {}
    for name in dir(config):
        if name.startswith("__"):
            continue
        value = getattr(config, name)
        if callable(value):
            continue
        if isinstance(value, (str, int, float, bool, list, tuple, dict, set)) or value is None:
            values[name] = _json_safe(value)
    return values


def _active_symbols(global_cfg):
    symbols = []
    raw = global_cfg.get("BOT_ACTIVE_SYMBOLS") or getattr(config, "BOT_ACTIVE_SYMBOLS", [])
    if isinstance(raw, (list, tuple, set)):
        for symbol in raw:
            symbol = str(symbol or "").strip().upper()
            if symbol and symbol not in symbols:
                symbols.append(symbol)
    return symbols


def _known_symbols(global_cfg):
    symbols = []
    for source in (
        getattr(config, "COIN_LIST", []),
        getattr(config, "SYMBOLS", []),
        global_cfg.get("BOT_ACTIVE_SYMBOLS", []),
    ):
        if isinstance(source, (list, tuple, set)):
            for symbol in source:
                symbol = str(symbol or "").strip().upper()
                if symbol and symbol not in symbols:
                    symbols.append(symbol)
    return symbols


def _add_symbol(out, symbol):
    symbol = str(symbol or "").strip().upper()
    if symbol:
        out.add(symbol)


def _symbols_from_live_signals(data):
    symbols = set()
    if not isinstance(data, dict):
        return symbols
    heartbeat = data.get("brain_heartbeat") or {}
    for symbol in heartbeat.get("active_symbols", []) if isinstance(heartbeat, dict) else []:
        _add_symbol(symbols, symbol)
    for sig in data.get("pending_signals", []) or []:
        if isinstance(sig, dict):
            _add_symbol(symbols, sig.get("symbol"))
    return symbols


def _symbols_from_state(data):
    symbols = set()
    if not isinstance(data, dict):
        return symbols

    def walk(value):
        if isinstance(value, dict):
            if "symbol" in value:
                _add_symbol(symbols, value.get("symbol"))
            for key in ("active_symbols", "BOT_ACTIVE_SYMBOLS", "symbols"):
                raw = value.get(key)
                if isinstance(raw, (list, tuple, set)):
                    for symbol in raw:
                        _add_symbol(symbols, symbol)
            for item in value.values():
                walk(item)
        elif isinstance(value, list):
            for item in value:
                walk(item)

    walk(data)
    return symbols


def _symbols_from_workbook(path):
    symbols = set()
    if not path or not os.path.exists(path):
        return symbols
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet_name in ("open_trades", "closed_trades"):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            rows = ws.iter_rows(values_only=True)
            headers = next(rows, None) or []
            try:
                symbol_idx = list(headers).index("Symbol")
            except ValueError:
                continue
            for row in rows:
                if row and symbol_idx < len(row):
                    _add_symbol(symbols, row[symbol_idx])
        try:
            wb.close()
        except Exception:
            pass
    except Exception:
        return symbols
    return symbols


def _relevant_symbols(global_cfg, raw_sources):
    symbols = set(_active_symbols(global_cfg))
    sources = raw_sources or {}
    live_data = ((sources.get("live_signals") or {}).get("data") or {})
    bot_state = ((sources.get("bot_state") or {}).get("data") or {})
    symbols.update(_symbols_from_live_signals(live_data))
    symbols.update(_symbols_from_state(bot_state))
    symbols.update(_symbols_from_workbook(paths.export_path()))
    symbols.update(_symbols_from_workbook(paths.history_path()))

    known = _known_symbols(global_cfg)
    known_set = set(known)
    if known_set:
        symbols = {symbol for symbol in symbols if symbol in known_set}

    ordered = []
    for symbol in known + sorted(symbols):
        if symbol in symbols and symbol not in ordered:
            ordered.append(symbol)
    return ordered


def _source_paths():
    import core.storage_manager as storage_manager

    return {
        "brain_settings": getattr(storage_manager, "BRAIN_FILE", None),
        "symbol_overrides": getattr(storage_manager, "SYMBOL_OVERRIDES_FILE", None),
        "tsl_settings": os.path.join(paths.account_dir(), "tsl_settings.json"),
        "presets_config": os.path.join(paths.account_dir(), "presets_config.json"),
        "grid_settings": os.path.join(paths.account_dir(), "grid_settings.json"),
        "hedge_settings": os.path.join(paths.account_dir(), "hedge_settings.json"),
        "bot_state": getattr(storage_manager, "STATE_FILE", None),
        "grid_state": os.path.join(paths.account_dir(), "grid_state.json"),
        "hedge_state": os.path.join(paths.account_dir(), "hedge_state.json"),
        "live_signals": os.path.join(paths.account_dir(), "live_signals.json"),
        "system_meta": getattr(storage_manager, "SYSTEM_META_FILE", None),
    }


def _stable_hash(payload):
    raw = json.dumps(_json_safe(payload), sort_keys=True, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


def _advisor_guide():
    return {
        "purpose": "Use this package to review RAT6 trading performance, risk controls, module behavior, and configuration drift. Do not propose direct automated order actions.",
        "files": {
            "technical_settings.json": "Current config snapshot plus raw source JSON files.",
            "advisor_export.xlsx": "Filtered trade/config/event view for the selected export window.",
            "user_context.md": "Human notes, goals, constraints, and questions from the operator.",
        },
        "timeframe_groups": {
            "G0": "Macro/base timeframe, usually highest timeframe.",
            "G1": "Trend/context timeframe.",
            "G2": "Execution/swing timeframe.",
            "G3": "Fast confirmation timeframe.",
        },
        "modules": {
            "TSL": "Trailing stop layer. Includes BE, PNL, STEP_R, SWING, BE_CASH, PSAR, ANTI_CASH.",
            "E/E": "Entry/Exit tactic layer. Includes fallback R, retest, structure, fib, pullback.",
            "DCA": "Adds to losing/averaging basket according to DCA rules.",
            "PCA": "Adds to winning/confirmed basket according to PCA rules.",
            "REV_C": "Recovery/reversal close logic.",
            "A.CUT": "Anti-cash hard stop/giveback guard.",
            "GRID": "Grid trading module.",
            "HEDGE": "Hedge trading module.",
            "SANDBOX": "Strategy sandbox/rules preview and bot strategy configuration.",
        },
        "trade_metrics": {
            "MAE": "Maximum adverse excursion in USD for a trade/session.",
            "MFE": "Maximum favorable excursion in USD for a trade/session.",
            "Fee": "Commission/spread fee estimate stored by the bot.",
            "Session ID": "Bot/manual session grouping key.",
        },
        "important_advice_rules": [
            "Prefer diagnosing which module/rule caused losses or missed profit.",
            "Compare performance by symbol, close reason, module tag, and trigger.",
            "Use current config values only as context; do not assume the bot should modify files automatically.",
        ],
    }


def build_snapshot(reason="manual"):
    import core.storage_manager as storage_manager

    paths.ensure_advisor_dirs()
    global_cfg = storage_manager.load_brain_settings()
    source_paths = _source_paths()
    raw_sources = {
        name: {"path": path, "data": _read_json_file(path)}
        for name, path in source_paths.items()
        if path
    }
    relevant_symbols = _relevant_symbols(global_cfg, raw_sources)
    omitted_symbols = [symbol for symbol in _known_symbols(global_cfg) if symbol not in set(relevant_symbols)]
    active_by_symbol = {}
    for symbol in relevant_symbols:
        try:
            active_by_symbol[symbol] = storage_manager.get_brain_settings_for_symbol(symbol)
        except Exception as exc:
            active_by_symbol[symbol] = {"_advisor_merge_error": str(exc)}

    config_payload = {
        "config_py": _public_config_values(),
        "active_global": _json_safe(global_cfg),
        "active_by_symbol": _json_safe(active_by_symbol),
        "relevant_symbols": relevant_symbols,
        "omitted_symbols": omitted_symbols,
        "raw_sources": _json_safe(raw_sources),
    }
    snapshot_id = _stable_hash(config_payload)

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "reason": reason,
        "account_id": paths.account_id(),
        "account_dir": paths.account_dir(),
        "config_snapshot_id": snapshot_id,
        "hash_basis": "config_py + active_global + active_by_symbol + raw_settings_sources",
        "advisor_guide": _advisor_guide(),
        "settings": config_payload,
    }


def build_technical_settings(reason="manual"):
    return build_snapshot(reason=reason)


def compact_snapshot_json(snapshot, limit=30000):
    raw = json.dumps(_json_safe(snapshot), ensure_ascii=False, sort_keys=True)
    if len(raw) <= limit:
        return raw
    return raw[:limit] + "...[truncated]"


def flatten_dict(data, prefix=""):
    out = {}
    if isinstance(data, dict):
        for key, value in data.items():
            next_key = f"{prefix}.{key}" if prefix else str(key)
            out.update(flatten_dict(value, next_key))
    elif isinstance(data, list):
        out[prefix] = json.dumps(_json_safe(data), ensure_ascii=False, sort_keys=True)
    else:
        out[prefix] = _json_safe(data)
    return out


def diff_snapshots(old_snapshot, new_snapshot, limit=300):
    old_flat = flatten_dict((old_snapshot or {}).get("settings", old_snapshot or {}))
    new_flat = flatten_dict((new_snapshot or {}).get("settings", new_snapshot or {}))
    changes = []
    for key in sorted(set(old_flat) | set(new_flat)):
        old_val = old_flat.get(key)
        new_val = new_flat.get(key)
        if old_val != new_val:
            changes.append((key, old_val, new_val))
            if len(changes) >= limit:
                changes.append(("_advisor_diff_truncated", "", f"limit={limit}"))
                break
    return changes


def clone_json(value):
    return copy.deepcopy(_json_safe(value))
