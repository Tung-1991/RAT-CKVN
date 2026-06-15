# -*- coding: utf-8 -*-
import csv
import json
import os
import time
import shutil
from datetime import datetime, timedelta

from . import config_snapshot, paths


EXPORT_SNAPSHOT_JSON_LIMIT = 6000
EXPORT_PAYLOAD_JSON_LIMIT = 4000


SHEETS = {
    "closed_trades": [
        "Recorded At", "Ticket", "Symbol", "Direction", "Lot", "Entry Time", "Exit Time",
        "Hold Seconds", "Entry Price", "Exit Price", "SL", "TP", "Fee", "Commission",
        "Swap", "Profit", "Close Reason", "Market Mode", "Trigger", "Session ID",
        "Signal Group", "Tactic", "Entry Exit Tactic", "Parent Ticket", "Source Type",
        "MAE", "MFE", "Module Tags", "Config Snapshot ID",
    ],
    "open_trades": [
        "Recorded At", "Ticket", "Symbol", "Direction", "Lot", "Open Time", "Entry Price",
        "SL", "TP", "Profit", "Swap", "Commission", "Tactic", "Entry Exit Tactic",
        "Market Mode", "MAE", "MFE", "Config Snapshot ID",
    ],
    "config_snapshots": [
        "Timestamp", "Snapshot ID", "Reason", "Account ID", "Snapshot JSON",
    ],
    "config_changes": [
        "Timestamp", "Old Snapshot ID", "New Snapshot ID", "Changed Path", "Old Value", "New Value",
    ],
    "events": [
        "Timestamp", "Severity", "Event Type", "Message", "Payload JSON",
    ],
    "summary_daily": ["Key", "Trades", "Profit", "Fee", "Wins", "Losses"],
    "summary_symbol": ["Key", "Trades", "Profit", "Fee", "Wins", "Losses"],
    "summary_timeframe": ["Key", "Trades", "Profit", "Fee", "Wins", "Losses"],
    "summary_signal_group": ["Key", "Trades", "Profit", "Fee", "Wins", "Losses"],
    "summary_close_reason": ["Key", "Trades", "Profit", "Fee", "Wins", "Losses"],
    "summary_module": ["Key", "Trades", "Profit", "Fee", "Wins", "Losses"],
    "trade_config_map": [
        "Recorded At", "Ticket", "Symbol", "Open Time", "Config Snapshot ID", "Source", "Payload JSON",
    ],
}


def _now():
    return datetime.now().isoformat(timespec="seconds")


def _json(value, limit=30000):
    raw = json.dumps(config_snapshot.clone_json(value), ensure_ascii=False, sort_keys=True)
    if len(raw) <= limit:
        return raw
    return raw[:limit] + "...[truncated]"


def _import_openpyxl():
    from openpyxl import Workbook, load_workbook

    return Workbook, load_workbook


def _load_workbook():
    Workbook, load_workbook = _import_openpyxl()
    paths.ensure_advisor_dirs()
    path = paths.history_path()
    for legacy in (paths.legacy_account_history_path(), paths.legacy_history_path()):
        if not os.path.exists(path) and os.path.exists(legacy):
            shutil.copy2(legacy, path)
    if os.path.exists(path):
        try:
            wb = load_workbook(path)
        except Exception:
            corrupt_path = f"{path}.corrupt_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            try:
                os.replace(path, corrupt_path)
            except Exception:
                pass
            wb = Workbook()
            default = wb.active
            wb.remove(default)
    else:
        wb = Workbook()
        default = wb.active
        wb.remove(default)
    for name, headers in SHEETS.items():
        if name not in wb.sheetnames:
            ws = wb.create_sheet(name)
            ws.append(headers)
        else:
            ws = wb[name]
            if ws.max_row == 0:
                ws.append(headers)
    return wb


def _save_workbook(wb):
    path = paths.history_path()
    tmp = f"{path}.tmp.xlsx"
    wb.save(tmp)
    os.replace(tmp, path)


def _load_export_workbook():
    Workbook, _load_workbook_impl = _import_openpyxl()
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for name, headers in SHEETS.items():
        ws = wb.create_sheet(name)
        ws.append(headers)
    return wb


def _save_export_workbook(wb):
    paths.ensure_advisor_dirs()
    path = paths.export_path()
    tmp = f"{path}.tmp.xlsx"
    wb.save(tmp)
    os.replace(tmp, path)


def _append_row(sheet_name, row):
    wb = _load_workbook()
    wb[sheet_name].append(row)
    _save_workbook(wb)


def record_event(event_type, message="", severity="INFO", payload=None):
    try:
        _append_row("events", [_now(), severity, event_type, message, _json(payload or {})])
        return True
    except Exception:
        return False


def _row_values(ws, row_idx):
    return [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]


def _header_index(headers, name):
    try:
        return headers.index(name)
    except ValueError:
        return None


def _truncate_export_value(value, limit):
    if value is None:
        return value
    text = str(value)
    if len(text) <= limit:
        return value
    return text[:limit] + "...[truncated_for_advisor_export]"


def _compact_export_row(sheet_name, headers, row_values):
    values = list(row_values)
    if sheet_name == "config_snapshots":
        pos = _header_index(headers, "Snapshot JSON")
        if pos is not None and pos < len(values):
            values[pos] = _truncate_export_value(values[pos], EXPORT_SNAPSHOT_JSON_LIMIT)
    elif sheet_name in {"events", "trade_config_map"}:
        pos = _header_index(headers, "Payload JSON")
        if pos is not None and pos < len(values):
            values[pos] = _truncate_export_value(values[pos], EXPORT_PAYLOAD_JSON_LIMIT)
    return values


def _config_ids_from_sheet(ws):
    if ws.max_row < 2:
        return set()
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    pos = _header_index(headers, "Config Snapshot ID")
    if pos is None:
        return set()
    ids = set()
    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row_idx, pos + 1).value
        if value:
            ids.add(str(value))
    return ids


def _latest_snapshot_from_sheet(ws):
    if ws.max_row < 2:
        return None
    raw = ws.cell(ws.max_row, 5).value
    if not raw or str(raw).endswith("[truncated]"):
        return None
    try:
        return json.loads(raw)
    except Exception:
        return None


def ensure_config_snapshot(reason="observer"):
    try:
        snapshot = config_snapshot.build_snapshot(reason=reason)
        snapshot_id = snapshot.get("config_snapshot_id")
        wb = _load_workbook()
        ws = wb["config_snapshots"]
        existing_ids = {str(ws.cell(r, 2).value) for r in range(2, ws.max_row + 1)}
        latest = _latest_snapshot_from_sheet(ws)
        latest_id = ws.cell(ws.max_row, 2).value if ws.max_row >= 2 else None
        if snapshot_id not in existing_ids:
            ws.append([
                snapshot.get("generated_at"),
                snapshot_id,
                reason,
                snapshot.get("account_id"),
                config_snapshot.compact_snapshot_json(snapshot),
            ])
            if latest_id and latest_id != snapshot_id:
                changes = config_snapshot.diff_snapshots(latest, snapshot) if latest else [
                    ("snapshot_id", latest_id, snapshot_id)
                ]
                ch_ws = wb["config_changes"]
                for path, old_val, new_val in changes:
                    ch_ws.append([
                        _now(),
                        latest_id,
                        snapshot_id,
                        path,
                        _json(old_val, limit=8000),
                        _json(new_val, limit=8000),
                    ])
        _save_workbook(wb)
        return snapshot_id
    except Exception as exc:
        record_event("config_snapshot_error", str(exc), severity="ERROR")
        return "unknown"


def _find_trade_snapshot(ticket):
    try:
        wb = _load_workbook()
        ws = wb["trade_config_map"]
        ticket_str = str(ticket)
        for row in range(ws.max_row, 1, -1):
            if str(ws.cell(row, 2).value) == ticket_str:
                return ws.cell(row, 5).value
    except Exception:
        pass
    return None


def record_trade_opened_data(ticket, symbol="", open_time="", payload=None, source="discovered"):
    try:
        ticket = str(ticket)
        if not ticket:
            return False
        snapshot_id = ensure_config_snapshot(reason=f"trade_open:{ticket}")
        wb = _load_workbook()
        ws = wb["trade_config_map"]
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row, 2).value) == ticket:
                return False
        if isinstance(open_time, (int, float)):
            open_time = datetime.fromtimestamp(open_time).isoformat(timespec="seconds")
        ws.append([_now(), ticket, symbol, open_time, snapshot_id, source, _json(payload)])
        _save_workbook(wb)
        return True
    except Exception as exc:
        record_event("trade_open_record_error", str(exc), severity="ERROR")
        return False


def record_trade_opened(pos, state=None, market_context=None, source="discovered"):
    ticket = str(getattr(pos, "ticket", ""))
    payload = {
        "magic": getattr(pos, "magic", None),
        "type": getattr(pos, "type", None),
        "volume": getattr(pos, "volume", None),
        "price_open": getattr(pos, "price_open", None),
        "sl": getattr(pos, "sl", None),
        "tp": getattr(pos, "tp", None),
        "state_tactic": (state or {}).get("trade_tactics", {}).get(ticket),
        "market_context": market_context or {},
    }
    return record_trade_opened_data(
        ticket,
        symbol=getattr(pos, "symbol", ""),
        open_time=getattr(pos, "time", ""),
        payload=payload,
        source=source,
    )


def _module_tags(reason, trigger, tactic, session_id):
    text = "|".join(str(v or "").upper() for v in [reason, trigger, tactic, session_id])
    tags = []
    for tag in ["REV_C", "BE_CASH", "TSL", "SL", "GRID", "HEDGE", "DCA", "PCA", "BE", "ANTI_CASH"]:
        if tag in text:
            tags.append(tag)
    return ",".join(tags) if tags else "unknown"


def _source_type(trigger, session_id):
    text = f"{trigger}|{session_id}".upper()
    if "HEDGE" in text:
        return "HEDGE"
    if "GRID" in text:
        return "GRID"
    if "[USER]" in text:
        return "MANUAL"
    if "[BOT]" in text or "AUTO" in text:
        return "BOT"
    return "unknown"


def _parse_signal_group(trigger):
    text = str(trigger or "").upper()
    for group in ["G0", "G1", "G2", "G3"]:
        if group in text:
            return group
    return "unknown"


def _safe_float(value, default=0.0):
    try:
        return float(str(value).replace("$", "").replace(",", "").strip())
    except Exception:
        return default


def _safe_datetime(value):
    if isinstance(value, datetime):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for candidate in (text, text.replace("Z", "+00:00")):
        try:
            parsed = datetime.fromisoformat(candidate)
            return parsed.replace(tzinfo=None) if parsed.tzinfo else parsed
        except Exception:
            pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            pass
    return None


def _derive_csv_trade_times(row_map):
    open_time = row_map.get("Open Time") or ""
    close_time = row_map.get("Close Time") or ""
    if _safe_datetime(close_time):
        return open_time, close_time
    session_id = str(row_map.get("Session_ID") or row_map.get("Session ID") or "")
    time_display = str(row_map.get("Time") or "")
    close_dt = None
    open_dt = _safe_datetime(open_time)
    try:
        if len(session_id) >= 8 and session_id[:8].isdigit():
            base_date = datetime.strptime(session_id[:8], "%Y%m%d").date()
            parts = time_display.split("->")
            open_part = parts[0].strip()
            close_part = parts[-1].strip()
            if len(close_part) >= 8:
                close_dt = datetime.combine(base_date, datetime.strptime(close_part[:8], "%H:%M:%S").time())
            if not open_dt and len(open_part) >= 8:
                open_dt = datetime.combine(base_date, datetime.strptime(open_part[:8], "%H:%M:%S").time())
    except Exception:
        pass
    if not close_dt:
        return (open_dt.isoformat(timespec="seconds") if open_dt else ""), ""
    return (
        open_dt.isoformat(timespec="seconds") if open_dt else "",
        close_dt.isoformat(timespec="seconds"),
    )


def _row_trade_datetime(headers, row_values):
    idx = {name: pos for pos, name in enumerate(headers)}
    exit_pos = idx.get("Exit Time")
    if exit_pos is not None and exit_pos < len(row_values):
        parsed = _safe_datetime(row_values[exit_pos])
        if parsed:
            return parsed

    entry_pos = idx.get("Entry Time")
    session_pos = idx.get("Session ID")
    entry_time = row_values[entry_pos] if entry_pos is not None and entry_pos < len(row_values) else ""
    session_id = str(row_values[session_pos] if session_pos is not None and session_pos < len(row_values) else "")
    has_session_date = len(session_id) >= 8 and session_id[:8].isdigit()
    if not entry_time and not has_session_date:
        return None

    recorded_pos = idx.get("Recorded At")
    if recorded_pos is not None and recorded_pos < len(row_values):
        return _safe_datetime(row_values[recorded_pos])
    return None


def _copy_sheet_rows(src_ws, dst_ws, include_row, transform_row=None):
    headers = [src_ws.cell(1, col).value for col in range(1, src_ws.max_column + 1)]
    copied = 0
    for row_idx in range(2, src_ws.max_row + 1):
        row_values = _row_values(src_ws, row_idx)
        if include_row(headers, row_values):
            if transform_row:
                row_values = transform_row(headers, row_values)
            dst_ws.append(row_values)
            copied += 1
    return copied


def record_closed_trade(
    ticket,
    symbol,
    direction,
    volume,
    entry_price,
    sl,
    tp,
    fee,
    pnl,
    close_reason,
    market_mode="ANY",
    trigger_signal="UNK",
    session_id="LEGACY",
    open_time_str="",
    mae_usd=0.0,
    mfe_usd=0.0,
    exit_time=None,
    state=None,
    warn_missing_snapshot=True,
):
    try:
        ticket_str = str(ticket)
        snapshot_id = _find_trade_snapshot(ticket_str)
        if not snapshot_id:
            snapshot_id = ensure_config_snapshot(reason=f"trade_close_missing_open_snapshot:{ticket_str}")
            if warn_missing_snapshot:
                record_event(
                    "missing_trade_open_snapshot",
                    f"No open-time config snapshot for ticket {ticket_str}; using current snapshot.",
                    severity="WARN",
                    payload={"ticket": ticket_str, "snapshot_id": snapshot_id},
                )

        state = state or {}
        tactic = state.get("trade_tactics", {}).get(ticket_str, "unknown")
        ee_tactic = state.get("entry_exit_tactics", {}).get(ticket_str, "unknown")
        parent = state.get("child_to_parent", {}).get(ticket_str, "")
        modules = _module_tags(close_reason, trigger_signal, tactic, session_id)
        if exit_time is None:
            exit_time = _now()
        hold_seconds = ""
        try:
            if open_time_str and exit_time:
                hold_seconds = max(0, int((datetime.fromisoformat(exit_time) - datetime.fromisoformat(open_time_str)).total_seconds()))
        except Exception:
            hold_seconds = ""

        row = [
            _now(),
            ticket_str,
            symbol,
            direction,
            volume,
            open_time_str or "",
            exit_time,
            hold_seconds,
            entry_price,
            "",
            sl,
            tp,
            fee,
            fee,
            "",
            pnl,
            close_reason,
            market_mode,
            trigger_signal,
            session_id,
            _parse_signal_group(trigger_signal),
            tactic,
            ee_tactic,
            parent,
            _source_type(trigger_signal, session_id),
            mae_usd,
            mfe_usd,
            modules,
            snapshot_id,
        ]

        wb = _load_workbook()
        ws = wb["closed_trades"]
        replaced = False
        for idx in range(2, ws.max_row + 1):
            if str(ws.cell(idx, 2).value) == ticket_str:
                for col, value in enumerate(row, start=1):
                    ws.cell(idx, col).value = value
                replaced = True
                break
        if not replaced:
            ws.append(row)
        _save_workbook(wb)
        return True
    except Exception as exc:
        record_event("closed_trade_record_error", str(exc), severity="ERROR", payload={"ticket": str(ticket)})
        return False


def sync_from_master_csv():
    try:
        import core.storage_manager as storage_manager

        csv_path = getattr(storage_manager, "MASTER_LOG_FILE", "")
        if not csv_path or not os.path.exists(csv_path):
            record_event("missing_master_trade_csv", "trade_history_master.csv not found", severity="WARN")
            return 0
        count = 0
        with open(csv_path, "r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            header = header or []
            for row in reader:
                if len(row) < 14:
                    continue
                row_map = {
                    name: row[idx] if idx < len(row) else ""
                    for idx, name in enumerate(header)
                }
                open_time, close_time = _derive_csv_trade_times(row_map)
                record_closed_trade(
                    row_map.get("Ticket", row[1] if len(row) > 1 else ""),
                    row_map.get("Symbol", row[2] if len(row) > 2 else ""),
                    row_map.get("Type", row[3] if len(row) > 3 else ""),
                    row_map.get("Vol", row[4] if len(row) > 4 else ""),
                    row_map.get("Entry", row[5] if len(row) > 5 else ""),
                    row_map.get("SL", row[6] if len(row) > 6 else ""),
                    row_map.get("TP", row[7] if len(row) > 7 else ""),
                    row_map.get("Fee", row[8] if len(row) > 8 else ""),
                    row_map.get("PnL ($)", row[9] if len(row) > 9 else ""),
                    row_map.get("Reason", row[10] if len(row) > 10 else ""),
                    market_mode=row_map.get("Market Mode", row[11] if len(row) > 11 else "ANY"),
                    trigger_signal=row_map.get("Trigger", row[12] if len(row) > 12 else "UNK"),
                    session_id=row_map.get("Session_ID", row[13] if len(row) > 13 else "LEGACY"),
                    open_time_str=open_time,
                    mae_usd=row_map.get("MAE ($)", row[14] if len(row) > 14 else 0.0),
                    mfe_usd=row_map.get("MFE ($)", row[15] if len(row) > 15 else 0.0),
                    exit_time=close_time,
                    warn_missing_snapshot=False,
                )
                count += 1
        return count
    except Exception as exc:
        record_event("master_csv_sync_error", str(exc), severity="ERROR")
        return 0


def refresh_open_trades(connector=None, state=None, market_contexts=None):
    try:
        wb = _load_workbook()
        ws = wb["open_trades"]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        positions = []
        if connector:
            try:
                positions = connector.get_all_open_positions() or []
            except Exception as exc:
                record_event("open_positions_read_error", str(exc), severity="WARN")
        state = state or {}
        for pos in positions:
            ticket = str(getattr(pos, "ticket", ""))
            snapshot_id = _find_trade_snapshot(ticket) or ensure_config_snapshot(reason=f"open_trade_export:{ticket}")
            direction = "BUY" if int(getattr(pos, "type", 0) or 0) == 0 else "SELL"
            open_time = getattr(pos, "time", "")
            if isinstance(open_time, (int, float)):
                open_time = datetime.fromtimestamp(open_time).isoformat(timespec="seconds")
            ctx = (market_contexts or {}).get(getattr(pos, "symbol", ""), {})
            excursion = state.get("trade_excursions", {}).get(ticket, {})
            ws.append([
                _now(),
                ticket,
                getattr(pos, "symbol", ""),
                direction,
                getattr(pos, "volume", ""),
                open_time,
                getattr(pos, "price_open", ""),
                getattr(pos, "sl", ""),
                getattr(pos, "tp", ""),
                getattr(pos, "profit", ""),
                getattr(pos, "swap", ""),
                getattr(pos, "commission", ""),
                state.get("trade_tactics", {}).get(ticket, "unknown"),
                state.get("entry_exit_tactics", {}).get(ticket, "unknown"),
                ctx.get("market_mode", "unknown") if isinstance(ctx, dict) else "unknown",
                excursion.get("mae_usd", ""),
                excursion.get("mfe_usd", ""),
                snapshot_id,
            ])
        _save_workbook(wb)
        return len(positions)
    except Exception as exc:
        record_event("open_trades_refresh_error", str(exc), severity="ERROR")
        return 0


def rebuild_summaries(wb=None, save=True, path_kind="history"):
    try:
        wb = wb or _load_workbook()
        closed = wb["closed_trades"]
        headers = [closed.cell(1, col).value for col in range(1, closed.max_column + 1)]
        idx = {name: pos + 1 for pos, name in enumerate(headers)}

        groups = {
            "summary_daily": {},
            "summary_symbol": {},
            "summary_timeframe": {},
            "summary_signal_group": {},
            "summary_close_reason": {},
            "summary_module": {},
        }

        def add(sheet, key, profit, fee):
            item = groups[sheet].setdefault(str(key or "unknown"), {"trades": 0, "profit": 0.0, "fee": 0.0, "wins": 0, "losses": 0})
            item["trades"] += 1
            item["profit"] += profit
            item["fee"] += fee
            if profit >= 0:
                item["wins"] += 1
            else:
                item["losses"] += 1

        for row in range(2, closed.max_row + 1):
            profit = _safe_float(closed.cell(row, idx.get("Profit", 16)).value)
            fee = _safe_float(closed.cell(row, idx.get("Fee", 13)).value)
            exit_time = str(closed.cell(row, idx.get("Exit Time", 7)).value or "")
            date_key = exit_time[:10] if len(exit_time) >= 10 else "unknown"
            add("summary_daily", date_key, profit, fee)
            add("summary_symbol", closed.cell(row, idx.get("Symbol", 3)).value, profit, fee)
            add("summary_timeframe", closed.cell(row, idx.get("Market Mode", 18)).value, profit, fee)
            add("summary_signal_group", closed.cell(row, idx.get("Signal Group", 21)).value, profit, fee)
            add("summary_close_reason", closed.cell(row, idx.get("Close Reason", 17)).value, profit, fee)
            tags = str(closed.cell(row, idx.get("Module Tags", 28)).value or "unknown").split(",")
            for tag in tags:
                add("summary_module", tag.strip() or "unknown", profit, fee)

        for sheet, data in groups.items():
            ws = wb[sheet]
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            for key, item in sorted(data.items()):
                ws.append([key, item["trades"], round(item["profit"], 2), round(item["fee"], 2), item["wins"], item["losses"]])
        if save:
            if path_kind == "export":
                _save_export_workbook(wb)
            else:
                _save_workbook(wb)
        return True
    except Exception as exc:
        record_event("summary_rebuild_error", str(exc), severity="ERROR")
        return False


def build_export_workbook(export_days=7):
    try:
        try:
            days = max(1, int(export_days or 7))
        except Exception:
            days = 7
        cutoff = datetime.now() - timedelta(days=days)
        source = _load_workbook()
        export_wb = _load_export_workbook()

        closed_src = source["closed_trades"]
        closed_dst = export_wb["closed_trades"]

        def include_closed(headers, row_values):
            trade_time = _row_trade_datetime(headers, row_values)
            return bool(trade_time and trade_time >= cutoff)

        closed_count = _copy_sheet_rows(
            closed_src,
            closed_dst,
            include_closed,
            lambda headers, row: _compact_export_row("closed_trades", headers, row),
        )

        if "open_trades" in source.sheetnames:
            _copy_sheet_rows(
                source["open_trades"],
                export_wb["open_trades"],
                lambda _h, _r: True,
                lambda headers, row: _compact_export_row("open_trades", headers, row),
            )

        used_config_ids = _config_ids_from_sheet(export_wb["closed_trades"]) | _config_ids_from_sheet(export_wb["open_trades"])
        latest_snapshot_id = None
        if "config_snapshots" in source.sheetnames and source["config_snapshots"].max_row >= 2:
            latest_snapshot_id = source["config_snapshots"].cell(source["config_snapshots"].max_row, 2).value
            if latest_snapshot_id:
                used_config_ids.add(str(latest_snapshot_id))

        if "trade_config_map" in source.sheetnames:
            def include_trade_map(headers, row_values):
                pos = _header_index(headers, "Config Snapshot ID")
                if pos is None or pos >= len(row_values):
                    return True
                return not used_config_ids or str(row_values[pos]) in used_config_ids

            _copy_sheet_rows(
                source["trade_config_map"],
                export_wb["trade_config_map"],
                include_trade_map,
                lambda headers, row: _compact_export_row("trade_config_map", headers, row),
            )
            used_config_ids |= _config_ids_from_sheet(export_wb["trade_config_map"])

        if "config_snapshots" in source.sheetnames:
            def include_snapshot(headers, row_values):
                pos = _header_index(headers, "Snapshot ID")
                if pos is None or pos >= len(row_values):
                    return True
                return not used_config_ids or str(row_values[pos]) in used_config_ids

            _copy_sheet_rows(
                source["config_snapshots"],
                export_wb["config_snapshots"],
                include_snapshot,
                lambda headers, row: _compact_export_row("config_snapshots", headers, row),
            )

        if "config_changes" in source.sheetnames:
            def include_change(headers, row_values):
                old_pos = _header_index(headers, "Old Snapshot ID")
                new_pos = _header_index(headers, "New Snapshot ID")
                old_id = row_values[old_pos] if old_pos is not None and old_pos < len(row_values) else ""
                new_id = row_values[new_pos] if new_pos is not None and new_pos < len(row_values) else ""
                return not used_config_ids or str(old_id) in used_config_ids or str(new_id) in used_config_ids

            _copy_sheet_rows(
                source["config_changes"],
                export_wb["config_changes"],
                include_change,
                lambda headers, row: _compact_export_row("config_changes", headers, row),
            )

        for sheet_name in ("events",):
            if sheet_name in source.sheetnames and sheet_name in export_wb.sheetnames:
                _copy_sheet_rows(
                    source[sheet_name],
                    export_wb[sheet_name],
                    lambda _h, _r: True,
                    lambda headers, row: _compact_export_row(sheet_name, headers, row),
                )

        rebuild_summaries(wb=export_wb, save=False, path_kind="export")
        _save_export_workbook(export_wb)
        return {"ok": True, "path": paths.export_path(), "closed_trades": closed_count, "export_days": days}
    except Exception as exc:
        record_event("advisor_export_build_error", str(exc), severity="ERROR", payload={"export_days": export_days})
        return {"ok": False, "error": str(exc), "path": paths.export_path(), "closed_trades": 0, "export_days": export_days}
